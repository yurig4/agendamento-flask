# [source: 133] Imports originais
from flask import Flask, render_template, request, jsonify, send_from_directory # Adicionado jsonify e send_from_directory
from werkzeug.utils import secure_filename
import os
import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError # Importar para tratar erros da API
import csv
from datetime import datetime
import smtplib
from email.message import EmailMessage
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter



app = Flask(__name__)
# [source: 133] Chave secreta é necessária para flash, mas não estritamente para jsonify
# Mantenha se usar flash em outras partes ou sessões
app.secret_key = 'segredo-super-seguro'

# [source: 133] Configurações
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB Permitido

# [source: 133] Credenciais e Calendar ID (Confirmado pelo usuário)
SERVICE_ACCOUNT_FILE = 'credenciais.json'
SCOPES = ['https://www.googleapis.com/auth/calendar']
CALENDAR_ID = 'yurig4@gmail.com'

# [source: 133] Garante que a pasta de uploads existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def agendamento():
    if request.method == 'POST':
        try:
            # --- Coletar dados do formulário ---
            # [source: 134] Campos de texto e data/hora
            nome = request.form['nomeProfessor']
            email = request.form['emailProfessor']
            disciplina = request.form['disciplina']
            turma = request.form['turma']
            assunto = request.form['assunto']
            data = request.form['data'] # YYYY-MM-DD
            inicio = request.form['horarioInicio'] # HH:MM
            fim = request.form['horarioFim'] # HH:MM

            # [source: 95] Coletar checkboxes (laboratorio) - Vem como lista se múltiplos selecionados
            laboratorios = request.form.getlist('laboratorio') # Use getlist para checkboxes
            laboratorios_str = ", ".join(laboratorios) if laboratorios else "Nenhum selecionado"

            # [source: 81-86, 92] Coletar textareas
            vidrarias_grupo = request.form.get('vidrariasGrupo', '')
            vidrarias_comuns = request.form.get('vidrariasComuns', '')
            solucoes = request.form.get('solucoes', '')
            reagentes = request.form.get('reagentes', '')
            equipamentos = request.form.get('equipamentos', '')
            observacoes = request.form.get('observacoes', '')
            num_alunos = request.form.get('numAlunos', 'N/A')
            num_grupos = request.form.get('numGrupos', 'N/A')


            # --- Montar descrição detalhada para o evento ---
            # [source: 135, 138] Descrição original + novos campos
            descricao = (
                f"Solicitante: {nome} ({email})\n"
                f"Disciplina: {disciplina}\nTurma: {turma}\nAssunto: {assunto}\n"
                f"Alunos: {num_alunos} | Grupos: {num_grupos}\n"
                f"Laboratório(s): {laboratorios_str}\n\n"
                f"--- Materiais ---\n"
                f"Vidrarias por Grupo: {vidrarias_grupo}\n"
                f"Vidrarias Comuns: {vidrarias_comuns}\n"
                f"Soluções: {solucoes}\n"
                f"Reagentes: {reagentes}\n"
                f"Equipamentos: {equipamentos}\n\n"
                f"--- Observações ---\n{observacoes}\n\n"
            )

            # --- Upload do arquivo ---
            caminho_arquivo_salvo = None
            # [source: 135] Upload do roteiro
            if 'roteiro' in request.files:
                roteiro = request.files['roteiro']
                if roteiro and roteiro.filename != '':
                    # Validar extensão (opcional, mas recomendado)
                    allowed_extensions = {'pdf', 'doc', 'docx'}
                    if '.' in roteiro.filename and roteiro.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                        filename = secure_filename(roteiro.filename)
                        caminho_arquivo_salvo = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        roteiro.save(caminho_arquivo_salvo)
                        # Adicionar link ou referência ao arquivo na descrição
                        # Para um link funcionar, você precisaria de uma rota para servir os arquivos
                        descricao += f"Roteiro Anexado: {filename} (Salvo no servidor)\n"
                        # Ou, se configurar uma rota para servir arquivos:
                        # descricao += f"Roteiro: http://seu-dominio/{app.config['UPLOAD_FOLDER']}/{filename}\n"
                    else:
                       # *** MODIFICAÇÃO AQUI ***: Retornar erro se extensão inválida
                       return jsonify({'success': False, 'message': 'Tipo de arquivo inválido. Use PDF, DOC ou DOCX.'}), 400
                else:
                    descricao += "Roteiro: Nenhum arquivo enviado.\n"
            else:
                 descricao += "Roteiro: Não incluído no formulário.\n"


            # --- Criar evento no Google Calendar ---
            # [source: 136-137] Autenticação e serviço
            credentials = service_account.Credentials.from_service_account_file(
                SERVICE_ACCOUNT_FILE, scopes=SCOPES)
            service = build('calendar', 'v3', credentials=credentials)

            # [source: 138-139] Montar horário no formato ISO 8601 com Timezone
            # Assume que o fuso horário do servidor ou onde o Flask roda é compatível com America/Sao_Paulo
            # Ou que a hora inserida já considera esse fuso
            start_dt = f"{data}T{inicio}:00" # Deixa sem offset, a API usa o timeZone
            end_dt = f"{data}T{fim}:00"

            # [source: 138-140] Definição do evento
            evento = {
                'summary': f"Aula Prática: {disciplina} - {turma}", # Título mais informativo
                'location': f"Laboratório(s): {laboratorios_str}", # Localização mais específica
                'description': descricao, # Descrição detalhada montada acima
                'start': {
                    'dateTime': start_dt,
                    'timeZone': 'America/Sao_Paulo', # Essencial para a API interpretar corretamente
                },
                'end': {
                    'dateTime': end_dt,
                    'timeZone': 'America/Sao_Paulo',
                },
                # [source: 140] Adiciona o professor como convidado (attendee)
                # Você pode adicionar outros emails fixos se necessário
                #'attendees': [{'email': email}],
                # Adicionar lembretes (opcional)
                # 'reminders': {
                #     'useDefault': False,
                #     'overrides': [
                #         {'method': 'email', 'minutes': 24 * 60}, # Email 1 dia antes
                #         {'method': 'popup', 'minutes': 60},      # Notificação 1 hora antes
                #     ],
                # },
            }

            # [source: 140] Inserir evento
            criado = service.events().insert(calendarId=CALENDAR_ID,
                                             body=evento,
                                             sendNotifications=False # Enviar convite para o email do professor
                                             ).execute()

            arquivo_excel = 'agendamentos.xlsx'

            # Se já existe, carregamos. Se não, criamos.
            if os.path.exists(arquivo_excel):
                wb = load_workbook(arquivo_excel)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                # Cabeçalho na primeira linha
                ws.append([
                    'Data de envio', 'Nome', 'Email', 'Disciplina', 'Turma', 'Assunto',
                    'Data da aula', 'Início', 'Fim', 'Descrição', 'Link do Evento'
                ])

            # Dados do agendamento
            linha = [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                nome, email, disciplina, turma, assunto, data, inicio, fim,
                descricao.replace('\n', ' '),
                criado.get("htmlLink")
            ]

            # Adiciona à próxima linha da planilha
            ws.append(linha)

            # Auto-ajustar largura das colunas (opcional)
            for i, coluna in enumerate(linha, 1):
                col_letter = get_column_letter(i)
                ws.column_dimensions[col_letter].width = max(len(str(coluna)) + 2, 15)

            # Salvar
            wb.save(arquivo_excel)            
                                                    
                            #ENVIAR EMAIL DE CONFIRMAÇÃO

            def enviar_email_confirmacao(destinatario, nome_prof, link_evento):
                msg = EmailMessage()
                msg['Subject'] = 'Confirmação de Agendamento de Aula Prática'
                msg['From'] = 'yurig4@gmail.com'     # <-- seu Gmail
                msg['To'] = destinatario

                corpo = f"""
            Olá, {nome_prof}!

            Seu agendamento foi registrado com sucesso.

            📅 Assunto: {assunto}
            📚 Disciplina: {disciplina}
            👥 Turma: {turma}
            📍 Data: {data}
            ⏰ Horário: {inicio} até {fim}

            🔗 Link do evento no Google Calendar:
            {link_evento}

            Se você não solicitou este agendamento, favor entrar em contato.

            Atenciosamente,
            Sistema de Agendamento
            """

                msg.set_content(corpo)

                try:
                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                        smtp.login('yurig4@gmail.com', '')  # <-- substitua
                        smtp.send_message(msg)
                    print('E-mail enviado com sucesso.')
                except Exception as e:
                    print('Erro ao enviar e-mail:', e)

            # Chamada da função:
            enviar_email_confirmacao(email, nome, criado.get("htmlLink"))

            # *** MODIFICAÇÃO AQUI ***: Retornar JSON de sucesso
            # [source: 141] Mensagem de sucesso
            link_evento = criado.get("htmlLink")
            return jsonify({
                'success': True,
                'message': f'Agendamento criado com sucesso! Veja no Google Calendar.',
                'eventLink': link_evento # Enviar o link de volta (opcional)
            }), 200 # Status HTTP 200 OK

        # Tratar erros específicos da API do Google
        except HttpError as error:
             print(f'Erro na API do Google Calendar: {error}')
             # Tentar extrair uma mensagem de erro mais amigável do Google
             try:
                 error_details = error.resp.json()
                 error_message = error_details.get('error', {}).get('message', str(error))
             except:
                 error_message = str(error)
             # *** MODIFICAÇÃO AQUI ***: Retornar JSON de erro da API
             return jsonify({'success': False, 'message': f'Erro ao criar evento no Google Calendar: {error_message}'}), 500

        # Tratar outros erros (ex: arquivo inválido, campos faltando, etc.)
        except Exception as e:
            print(f'Erro inesperado: {e}')
            import traceback
            traceback.print_exc() # Imprimir stack trace completo no console do Flask para debug
            # *** MODIFICAÇÃO AQUI ***: Retornar JSON de erro genérico
            return jsonify({'success': False, 'message': f'Erro interno no servidor: {e}'}), 500

    # Se for método GET, apenas renderiza o template
    # [source: 141] Renderizar template no GET
    return render_template('agendamento.html')


# --- Rota Opcional para Servir Arquivos ---
# Se você quiser que o link do arquivo na descrição do evento funcione,
# você precisa de uma rota como esta. Ajuste a segurança conforme necessário.
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except FileNotFoundError:
        return "Arquivo não encontrado", 404


if __name__ == '__main__':
    # [source: 141] Executar app
    # Mude host='0.0.0.0' se quiser acessar de outros dispositivos na mesma rede
    app.run(debug=True, host='127.0.0.1', port=5000)